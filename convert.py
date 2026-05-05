"""
convert.py — Scoreboard Test.xlsx → output.json

Reads the clinic performance scoreboard and emits a clean, queryable JSON
file with one record per weekly row, plus a metadata block that documents
every metric's label, focus area, data source, and responsible role.

Usage
-----
    python3 convert.py [options]

Sort options
------------
    --sort date                       Sort records by week (default)
    --sort <metric_key>               Sort records by any numeric metric
    --order asc|desc                  Sort direction (default: asc)
    --sort-keys                       Sort metric keys A→Z within each record
    --output <path>                   Write to a custom file instead of output.json
    --list-metrics                    Print all available metric keys and exit

Examples
--------
    python3 convert.py
    python3 convert.py --sort total_revenue_all_services --order desc
    python3 convert.py --sort date --order desc --output newest_first.json
    python3 convert.py --sort-keys --output alpha_keys.json
    python3 convert.py --list-metrics
"""

import argparse
import json
import re
import sys
import datetime
from pathlib import Path
import openpyxl

INPUT_FILE = Path("Scoreboard Test.xlsx")
OUTPUT_FILE = Path("output.json")

# Row numbers (1-based, matching what you see in Excel)
ROW_SECTION_HEADER = 1   # merged-cell section banners (e.g. "PHONE PERFORMANCE")
ROW_METRIC_LABELS  = 2   # main column headers
ROW_FOCUS          = 3   # "Financial", "Marketing", …
ROW_SOURCE         = 4   # "EMR", "CallHero", …
ROW_ROLE           = 5   # "J", "Nicole", "Beth", …
ROW_TARGET_FLAGS   = 6   # "Target" marker
ROW_TARGET_VALUES  = 7   # spot target values / formula notes (col A == 6.0 here)
FIRST_DATA_ROW     = 8   # weekly records start here


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

def slugify(text: str) -> str:
    """Convert a header label to a safe, lowercase JSON key.

    Replaces % and # with readable words first so that "NAR % Collected"
    and "NAR Collected" don't collapse to the same slug. Then collapses
    runs of non-alphanumeric characters to a single underscore.
    """
    text = str(text).strip().replace("\n", " ")
    text = text.replace("%", "_pct").replace("#", "_num")
    text = re.sub(r"[^a-zA-Z0-9]+", "_", text)
    return text.strip("_").lower()


def make_unique_key(base_key: str, col_letter: str, seen: dict[str, str]) -> str:
    """Return a collision-free key, appending the column letter when needed.

    Using the column letter (e.g. "pva_4_wk_avg_df") is more traceable than
    a bare counter ("pva_4_wk_avg_2") — a developer can immediately look the
    column up in the original spreadsheet or in metadata["metrics"].
    """
    if base_key not in seen:
        seen[base_key] = col_letter
        return base_key
    # Collision: suffix with the column letter so the key stays traceable.
    return f"{base_key}_{col_letter.lower()}"


def coerce_value(raw):
    """Turn a raw cell value into a clean Python type.

    - datetime  → ISO-format string
    - int/float → kept as number
    - string    → stripped; numeric strings become float; "#REF!" / formula
                  errors become None
    - None      → None
    """
    if raw is None:
        return None

    if isinstance(raw, datetime.datetime):
        return raw.date().isoformat()

    if isinstance(raw, (int, float)):
        return raw

    text = str(raw).strip()

    if not text:
        return None

    # Formula errors produced by openpyxl when data_only=True
    if text.startswith("#") and text.endswith("!"):
        return None

    # Numbers stored as strings (e.g. "2.87", "3.26")
    try:
        return float(text)
    except ValueError:
        pass

    return text


def cell_val(ws, row: int, col: int):
    """Return the coerced value of a single cell."""
    return coerce_value(ws.cell(row=row, column=col).value)


# ---------------------------------------------------------------------------
# Spreadsheet structure parsers
# ---------------------------------------------------------------------------

def parse_section_headers(ws) -> dict[int, str]:
    """Map column index → section name from the merged-cell banner row.

    Row 1 may contain merged cells that label groups of columns
    (e.g. "PHONE PERFORMANCE" spanning AJ:AP). We expand each merged
    range so every covered column gets the banner text.
    """
    section_map: dict[int, str] = {}

    for merge in ws.merged_cells.ranges:
        if merge.min_row == ROW_SECTION_HEADER:
            value = ws.cell(merge.min_row, merge.min_col).value
            if value:
                for col in range(merge.min_col, merge.max_col + 1):
                    section_map[col] = str(value).strip()

    # Also pick up any un-merged values in row 1
    for cell in ws[ROW_SECTION_HEADER]:
        if cell.value and cell.column not in section_map:
            section_map[cell.column] = str(cell.value).strip()

    return section_map


def build_column_map(ws, section_map: dict[int, str]) -> dict[int, dict]:
    """Build a mapping from column index → metric descriptor dict.

    Skips columns that have no header label (spacer/blank columns).
    Column A (index 1) is the date column — handled separately in extract_records.
    """
    seen_keys: dict[str, str] = {}  # base_key → first column letter seen
    columns: dict[int, dict] = {}

    for cell in ws[ROW_METRIC_LABELS]:
        col_idx = cell.column

        if col_idx == 1 or cell.value is None:
            continue

        raw_label = str(cell.value).strip().replace("\n", " ")
        base_key  = slugify(raw_label)

        if not base_key:
            continue  # genuinely blank header → spacer column

        unique_key = make_unique_key(base_key, cell.column_letter, seen_keys)

        columns[col_idx] = {
            "key":     unique_key,
            "label":   raw_label,
            "column":  cell.column_letter,
            "section": section_map.get(col_idx),
            "focus":   cell_val(ws, ROW_FOCUS,         col_idx),
            "source":  cell_val(ws, ROW_SOURCE,        col_idx),
            "role":    cell_val(ws, ROW_ROLE,           col_idx),
            "target":  cell_val(ws, ROW_TARGET_VALUES,  col_idx),
        }

    return columns


def extract_records(ws, col_map: dict[int, dict]) -> list[dict]:
    """Extract each weekly data row as a flat dict keyed by metric slug.

    Rows where column A is not a datetime are skipped (e.g. the target-value
    row which has 6.0 in col A).
    Always returns records in chronological order; callers may re-sort afterwards.
    """
    records = []

    for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
        raw_date = row[0].value  # column A is always index 0 in the row tuple

        if not isinstance(raw_date, datetime.datetime):
            continue

        record: dict = {"week_ending": raw_date.date().isoformat()}

        for col_idx, meta in col_map.items():
            cell = ws.cell(row=row[0].row, column=col_idx)
            record[meta["key"]] = coerce_value(cell.value)

        records.append(record)

    records.sort(key=lambda r: r["week_ending"])
    return records


# ---------------------------------------------------------------------------
# Sorting functions
# ---------------------------------------------------------------------------

def sort_by_date(records: list[dict], order: str) -> list[dict]:
    """Sort records chronologically.

    order: 'asc'  → oldest week first (default extract order)
           'desc' → newest week first
    """
    reverse = order == "desc"
    return sorted(records, key=lambda r: r["week_ending"], reverse=reverse)


def sort_by_metric(records: list[dict], metric_key: str, order: str) -> list[dict]:
    """Sort records by a numeric metric field.

    Records where the metric is None are placed at the end regardless of order,
    so null-heavy columns don't crowd the top or bottom.

    order: 'asc'  → smallest value first (min → max)
           'desc' → largest value first  (max → min)
    """
    if not records:
        return records

    if metric_key not in records[0]:
        available = ", ".join(k for k in records[0] if k != "week_ending")
        raise ValueError(
            f"Unknown metric key: '{metric_key}'.\n"
            f"Run --list-metrics to see all available keys."
        )

    reverse = order == "desc"

    def sort_key(record):
        val = record.get(metric_key)
        # Push None values to the very end (use a tuple so None sorts after numbers)
        if val is None:
            return (1, 0)
        return (0, val if reverse is False else -val)

    return sorted(records, key=sort_key)


def sort_keys_alphabetically(records: list[dict]) -> list[dict]:
    """Re-order keys within each record alphabetically.

    'week_ending' is always kept first as the record identifier.
    Useful when scanning the JSON directly or diffing two outputs.
    """
    sorted_records = []
    for record in records:
        # Pull out week_ending, sort everything else, then reassemble
        other_keys = sorted(k for k in record if k != "week_ending")
        ordered = {"week_ending": record["week_ending"]}
        ordered.update({k: record[k] for k in other_keys})
        sorted_records.append(ordered)
    return sorted_records


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Convert Scoreboard Test.xlsx to a clean JSON file.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--sort",
        default="date",
        metavar="KEY",
        help=(
            "Field to sort records by. Use 'date' for chronological order "
            "or any metric key (e.g. total_revenue_all_services). "
            "Default: date"
        ),
    )
    parser.add_argument(
        "--order",
        choices=["asc", "desc"],
        default="asc",
        help="Sort direction: asc (min→max / oldest→newest) or desc (max→min / newest→oldest). Default: asc",
    )
    parser.add_argument(
        "--sort-keys",
        action="store_true",
        help="Sort metric keys A→Z within each record (week_ending always stays first).",
    )
    parser.add_argument(
        "--output",
        default=str(OUTPUT_FILE),
        metavar="PATH",
        help=f"Output file path. Default: {OUTPUT_FILE}",
    )
    parser.add_argument(
        "--list-metrics",
        action="store_true",
        help="Print all available metric keys and exit without writing JSON.",
    )
    return parser.parse_args()


def main():
    args = parse_args()

    if not INPUT_FILE.exists():
        raise FileNotFoundError(f"Input file not found: {INPUT_FILE}")

    # data_only=True makes openpyxl return cached formula results, not formulas
    wb = openpyxl.load_workbook(INPUT_FILE, data_only=True)
    ws = wb["SCOREBOARD"]

    section_map = parse_section_headers(ws)
    col_map     = build_column_map(ws, section_map)

    metrics_meta = {
        meta["key"]: {
            "label":   meta["label"],
            "column":  meta["column"],
            "section": meta["section"],
            "focus":   meta["focus"],
            "source":  meta["source"],
            "role":    meta["role"],
            "target":  meta["target"],
        }
        for meta in col_map.values()
    }

    # --list-metrics: print keys and exit without touching the output file
    if args.list_metrics:
        print(f"{'KEY':<45}  {'LABEL'}")
        print("-" * 80)
        for key, meta in metrics_meta.items():
            print(f"{key:<45}  {meta['label']}")
        sys.exit(0)

    records = extract_records(ws, col_map)

    # Apply sorting
    if args.sort == "date":
        records = sort_by_date(records, args.order)
    else:
        records = sort_by_metric(records, args.sort, args.order)

    if args.sort_keys:
        records = sort_keys_alphabetically(records)

    output = {
        "metadata": {
            "source_file":   INPUT_FILE.name,
            "sheet":         ws.title,
            "extracted_at":  datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "record_count":  len(records),
            "metric_count":  len(metrics_meta),
            "sort_applied": {
                "sort_by":    args.sort,
                "order":      args.order,
                "keys_alpha": args.sort_keys,
            },
            "metrics":       metrics_meta,
        },
        "records": records,
    }

    out_path = Path(args.output)
    out_path.write_text(json.dumps(output, indent=2, ensure_ascii=False))
    print(f"Written {len(records)} records, {len(metrics_meta)} metrics → {out_path}")


if __name__ == "__main__":
    main()
